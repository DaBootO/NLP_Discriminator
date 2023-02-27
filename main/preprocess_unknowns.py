import openpyxl
from openpyxl.utils import get_column_interval
import re
import pandas as pd
import os
import string
import time

def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

items = 12
letters = string.ascii_uppercase

directory_base_data = os.path.abspath(os.path.realpath(__file__)[:-len(os.path.basename(__file__))] + "../base_data/to_check/") + '/'

DF_list = []

for file in os.listdir(directory_base_data):
    if file.split('.')[-1] == "xlsx":
        print(file, "is being processed...")
        
        # I/O
        wb = openpyxl.load_workbook(filename=directory_base_data+str(file)) # type: (wb) -> openpyxl.
        ws = wb.active
        rows = len(tuple(ws.rows))
        columns = len(tuple(ws.columns))
        
        # find start of data
        row_start = 1
        while ws['C'+str(row_start)].value != "title":
            row_start += 1
        row_start += 1
        
        temp_df = load_workbook_range(f"A{row_start}:{letters[columns-1]}{rows}", ws)
        
        DF_list.append(temp_df)

Data = pd.concat(DF_list)

column_names = [
    "0/1",
    "comment",
    "title",
    "author",
    "url",
    "year",
    "num_citations",
    "num_versions",
    "cluster_id",
    "url_pdf",
    "url_citations",
    "url_versions",
    "url_citation",
    "excerpt"
]

col_dict = {string.ascii_uppercase[i]: column_names[i] for i in range(14)}

Data.rename(col_dict, axis=1, inplace=True)
Data.reset_index(drop=True, inplace=True)

Data.to_json(os.path.join(directory_base_data, "data_checks.json"))
Data.to_json(os.path.join(directory_base_data, f"data_checks{time.strftime('%H%M%S_%d%m%y')}.json"))
